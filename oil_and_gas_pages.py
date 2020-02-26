import selenium
from selenium import webdriver
from time import sleep
import openpyxl

category_name = "Abrasives"

# now i'm trying the abrasives page under classified section.
# hence this url
# url = "https://www.oilandgaspages.com/classilist.php?c=003&name=ABRASIVES"

# url = 'https://www.oilandgaspages.com/classifications.php'
driver = webdriver.Chrome(
    executable_path='C:\\Users\\Sohail21400\\Downloads\\chromedriver_win32\\chromedriver.exe')

driver.get(url)

sleep(7)

work_book = openpyxl.load_workbook('data.xlsx')
print("work book opened")
# # gets the list of all sheets in the workbook
list_of_sheets = work_book.sheetnames
list_of_links_sheet = work_book["Link of Categories"]

# if the sheet with the category name is present it takes it or else it will create one.
if category_name in list_of_sheets:
    sheet = work_book[category_name]
    print(f"sheet named {category_name} is already present")
else:
    sheet = work_book.create_sheet(category_name)
    print(f"sheet named {category_name} has created")

# ---------------------------------------------------------------------------

# this code is to know from where the program has started writing
# this will only write once the program is started
sheet.cell(sheet.max_row + 1, 1).value = "---------"


def is_next_page_available():
    try:
        nxt_button = driver.find_element_by_xpath('//*[@id="content"]/nav/ul[3]/li/a')
        next_page_link = nxt_button.get_attribute('href')
        # print(next_page_link)
        if next_page_link is not None:
            return True
        else:
            return False

    # if there is no next page there will be only one page
    except selenium.common.exceptions.NoSuchElementException:
        print("this website has only one page!")
        return False


def get_the_list_of_links():

    print("generating the list of links")
    list_of_links = []
    is_nxt_page_available = is_next_page_available()

    while is_nxt_page_available:

        title_links = driver.find_elements_by_xpath('//*[@id="content"]/div/a')

        for i in title_links:
            list_of_links.append(i.get_attribute('href'))

        is_nxt_page_available = is_next_page_available()

        if is_nxt_page_available:
            next_button = driver.find_element_by_xpath('//*[@id="content"]/nav/ul[3]/li/a')
            next_button.click()
            sleep(4)
    print("list of links generated...")
    return list_of_links


def get_main_details(link, row):
    driver.get(link)
    sleep(4)
    try:
        title = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div/div[2]/h1').text
    except selenium.common.exceptions.NoSuchElementException:
        title = "_"
    try:
        address = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div/div[2]/p').text
    except selenium.common.exceptions.NoSuchElementException:
        address = "_"
    try:
        phone = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div/div[2]/div/ul/li[1]/div').text
    except selenium.common.exceptions.NoSuchElementException:
        phone = "_"
    try:
        fax = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div/div[2]/div/ul/li[2]').text
    except selenium.common.exceptions.NoSuchElementException:
        fax = "_"
    try:
        website = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div/div[2]/div/ul/li[4]/div/a')\
            .get_attribute('href')
    except selenium.common.exceptions.NoSuchElementException:
        website = "_"
    try:
        description = driver.find_element_by_xpath('//*[@id="content"]/div[3]/p').text
    except selenium.common.exceptions.NoSuchElementException:
        description = "_"

    # print(title)
    # print(address)
    # print(phone)
    # print(fax)
    # print(website)
    # print(link)

    sheet.cell(row, 1).value = title
    sheet.cell(row, 2).value = address
    sheet.cell(row, 3).value = phone
    sheet.cell(row, 4).value = fax
    sheet.cell(row, 5).value = website
    sheet.cell(row, 6).value = description
    sheet.cell(row, 7).value = link
    print('main details have been saved........')


def get_executive_details(row):

    executive_info_box_xpath = '//*[@id="content"]/div[3]/div/div'

    executive_info_boxes = driver.find_elements_by_xpath(executive_info_box_xpath)
    number_of_executives = len(executive_info_boxes)

    # so we got the number of executives in the website

    # i is the number in the xpath
    # there is no reason why it's 2
    executive_number = 2
    gap = 0

    while executive_number < number_of_executives + 2:
        # if i = 1 it gives the details of first executive and so on

        # this for loop is for getting all the details
        # like phone number mail etc..
        for j in range(1, 5):
            # j == 1 --> name
            # j == 2 --> designation
            # j == 3 --> phone number
            # j == 4 --> email

            # is j == 4 it is the section of mail and it is given in the a tag
            # so we need to access the text inside it
            if j == 4:
                temp_xpath = f'//*[@id="content"]/div[3]/div[{executive_number}]/div/div[{j}]/p/a'
                try:
                    data = driver.find_element_by_xpath(temp_xpath).text
                except selenium.common.exceptions.NoSuchElementException:
                    data = "_"
                if data is "":
                    data = "_"

                # i'm directly setting the column 10 because this code will only execute when j == 4
                sheet.cell(row, gap + 11).value = data
            else:
                temp_xpath = f'//*[@id="content"]/div[3]/div[{executive_number}]/div/div[{j}]/p'
                try:
                    data = driver.find_element_by_xpath(temp_xpath).text
                except selenium.common.exceptions.NoSuchElementException:
                    data = "_"
                if data is "":
                    data = "_"
                sheet.cell(row, gap + j + 7).value = data

            # print(data)
            # print(f'{j} executive detail have been saved ')

        gap += 4
        executive_number += 1

# categories = list_of_links_sheet
# for one_category in range(1, categories):
#     
links = get_the_list_of_links()
number_of_links = len(links)
progress = 0

for a_link in links:
    last_row = sheet.max_row + 1
    get_main_details(a_link, last_row)
    get_executive_details(last_row)
    progress += 1
    print(f"progress = {progress}/{number_of_links} ")
    work_book.save("data.xlsx")
    print("workbook saved!")


# def get_category_links():
#     category_link_xpath = '//*[@id="results"]/tbody/tr/td/div/a'
#     category_links = driver.find_elements_by_xpath(category_link_xpath)
#
#     if "Link of Categories" in list_of_sheets:
#         sheet_of_category_links = work_book["Link of Categories"]
#         print(f"sheet named 'Link of Categories' is already present")
#     else:
#         sheet_of_category_links = work_book.create_sheet("Link of Categories")
#         print(f"sheet named 'Link of Categories'' has created")
#
#     for category_link in category_links:
#         data = category_link.get_attribute('href')
#         max_row = sheet_of_category_links.max_row
#         sheet_of_category_links.cell(max_row + 1, 1).value = data

